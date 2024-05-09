import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import pandas as pd
import sys
import yaml

# Get the parent directory of the current script's directory
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
project_parent = os.path.dirname(project_root)
# Add the parent directory of ai-capstone to the Python path
bebe_dir = os.path.join(project_parent, 'BEBE')
sys.path.append(bebe_dir)

from BEBE.visualization.time_series import plot_track

from utils.data_config import beh_names, beh_dict
from utils.visualization import plot_kill_predictions

# $ python scripts/check_predicted_kills.py --prediction_dir=/home/matthew/AI_Capstone/ai-capstone/BEBE_results_phase2/harnet/F202_15sample_16hz_2hr_harnet/F202_15sample_16hz_2hr_harnet/fold_1/ --output_dir=/home/matthew/AI_Capstone/ai-capstone/BEBE_results_phase2/post_process/ --config_dir=/home/matthew/AI_Capstone/ai-capstone/BEBE-datasets-phase2/F202_15sample_16hz_2hr/FormatData/

summary = """
Input - csv containing labeled behavior (can be ground truth or predictions)
Outputs - 

    CSV with false kills differentiated from real kills. 
        A kill can be marked false if any of the following conditions are met:
        -There is no stalking within MIN_STALK_DELAY of the kill
        -The stalking before the kill is less than MIN_STALK_TIME
        -The kill is shorter than MIN_KILL_TIME
        -(NOT IMPLEMENTED, is this needed?) The kill is longer than MAX_KILL_TIME
        -There is no feeding within MAX_FEED_DELTA
        -The feeding is shorter than MIN_FEED_TIME
    Plot:
        x, y, z, accel data
        input data (each time segment has a behavior assigned)
        updated data (false kills are labelled as a different behavior)


    Summary printout:
        Likely number of kills identified (max of 1 per each csv)
        Potential number of kills identified (no limit per csv)
        False kills removed (number of kills that violated rules above)

"""

# Constants
SAMPLING_RATE = 16  # Sampling rate in Hz
MIN_STALK_DELAY = 5 * SAMPLING_RATE  # Minimum delay between stalk end and kill start in samples (2 seconds at 16Hz) (allows for human labelling error)
MIN_STALK_TIME = 5 * SAMPLING_RATE  # Minimum stalking time in samples (4 seconds at 16Hz)
MIN_KILL_TIME = 3 * SAMPLING_RATE  # Minimum kill time in samples (2 seconds at 16Hz) (phase 1)
MAX_KILL_TIME = 15 * 60 * SAMPLING_RATE  # Maximum kill time in samples (15 minutes)

MIN_PHASE_2_TIME = 3 * SAMPLING_RATE
# NOT NEEDED MAX_KILL_TIME = 15 * 60 * SAMPLING_RATE  # Maximum kill time in samples (15 minutes)
MIN_FEED_DELTA = 60 * 60 * SAMPLING_RATE  # Minimum delay between end of kill and feed START in samples (15 minutes) (delay between end of PHASE 2 and start of FEED)
MAX_FEED_DELTA = 3 * 60 * 60 * SAMPLING_RATE # max amount of time between to look for FEED behavior after end of KILL
MIN_FEED_TIME = 5 * 60 * SAMPLING_RATE  # Minimum feeding duration in samples (2 minutes)


def check_kill_status(df, start_time, min_stalk_time=MIN_STALK_TIME, min_stalk_delay=MIN_STALK_DELAY,
                      min_kill_time=MIN_KILL_TIME, min_feed_delta=MIN_FEED_DELTA, max_feed_delta=MAX_FEED_DELTA,
                      min_feed_time=MIN_FEED_TIME):
    """
    Check if a kill behavior is valid based on the criteria at the top of the file

    :param df:
    :param start_time:
    :param min_duration:
    :return: True if all criteria are met, else False
    """

    # Check if start_time is within the index range of the DataFrame
    if start_time < 0 or start_time >= len(df):
        return False, "Start time out of range"

    # Check if the behavior at start_time is 'KILL'
    if df.at[start_time, 'behavior'] != beh_dict['KILL']:
        return False, "Behavior is not KILL"

    # Check if there is enough STALK behavior before the kill start
    stalk_count = 0
    stalk_start_window = start_time - min_stalk_time - min_stalk_delay - 1
    stalk_start_window = max(0, stalk_start_window)
    for i in range(start_time - 1, stalk_start_window, -1):
        if df.at[i, 'behavior'] == beh_dict['STALK']:
            stalk_count += 1
            if stalk_count >= min_stalk_time:
                break
        elif min_stalk_time <= stalk_count <= min_stalk_time + min_stalk_delay:
            break
        else:
            stalk_count = 0

    # either not enough stalk time OR the stalking occurred outside of the MIN_STALK_DELAY window
    if stalk_count < min_stalk_time:# or stalk_count < min_stalk_time + min_stalk_delay:
        return False, f"No stalking within MIN_STALK_DELAY ({min_stalk_delay/SAMPLING_RATE} sec)"

    # Check if the kill behavior lasts for at least min_duration time steps
    end_time = start_time + min_kill_time - 1
    if end_time >= len(df):
        return False, "End time out of range"

    # ensure the kill is long enough
    for i in range(start_time + 1, end_time + 1):
        if df.at[i, 'behavior'] != beh_dict['KILL']:
            kill_length = (i - start_time) / SAMPLING_RATE
            return False, f"Kill is not long enough ({kill_length} < {MIN_KILL_TIME/SAMPLING_RATE})"

    # find the actual end time of the kill
    actual_kill_end_time = start_time
    for i in range(start_time, len(df)):
        if df.at[i, 'behavior'] != beh_dict['KILL']:
            actual_kill_end_time = i - 1
            break

    # Check if there is FEED behavior within min_feed_delta of the end of the KILL behavior
    feed_time_check_end = min(actual_kill_end_time + max_feed_delta + 1, len(df))
    feed_start_min = actual_kill_end_time + min_feed_delta
    feed_count = 0
    for i in range(actual_kill_end_time, feed_time_check_end):
        if df.at[i, 'behavior'] == beh_dict['FEED']:
            feed_count += 1
        elif feed_count == 0 and i >= feed_start_min:
            # feeding has started early enough
            return False, "Feeding has not started within MIN_FEED_DELTA"

    # check for enough feed time
    if feed_count <= min_feed_time:
        feed_time = feed_count / SAMPLING_RATE
        return False, f"Feeding was shorter than MIN_FEED_TIME ({feed_time} < {MIN_FEED_TIME})"

    # All criteria met, return True
    return True, "All criteria met successfully"

def filter_model_predictions(prediction_csv, filtered_csv):
    # Load the CSV file into a DataFrame
    df = pd.read_csv(prediction_csv, header=None, names=['behavior'])

    # TODO: get rid of (WIP now)
    df[['behavior']].to_csv(filtered_csv, header=False, index=False)

    # Find the indices where 'kill' behavior starts
    kill_starts = df[df['behavior'] == beh_dict['KILL']].index

    # Find the indices where the behavior changes from any other behavior to 'kill'
    start_of_kill_sequences = [index for index in kill_starts if
                               index == 0 or df.at[index - 1, 'behavior'] != beh_dict['KILL']]

    kill_statuses = {}
    print("Start of kill sequences:", start_of_kill_sequences)
    for start in start_of_kill_sequences:
        valid_kill, message = check_kill_status(df, start)
        kill_statuses[start] = valid_kill
        print(f"CHECK {start=}, {valid_kill=}, {message=}")


    # TODO: do we need to change kill labels?
    fname = "blah.png"
    plot_kill_predictions(df, kill_statuses, fname)
    print("DONE")
    
    # Save the updated DataFrame to a new CSV file
    # df.to_csv(filtered_csv, header=False, index=False)
    df[['behavior']].to_csv(filtered_csv, header=False, index=False)

    return prediction_csv, kill_statuses

# Function to calculate the duration of behaviors
def calculate_duration(behavior_window, behavior_type):
    durations = []
    duration = 0
    for behavior in behavior_window:
        if behavior == behavior_type:
            duration += 1
        else:
            if duration > 0:
                durations.append(duration)
            duration = 0
    if duration > 0:
        durations.append(duration)
    return durations


def create_spreadsheet(prediction_csv, filtered_csv, output_dir):
    # Define behavior names and their corresponding indices
    beh_names = ['unknown', 'STALK', 'KILL', 'KILL_PHASE2', 'FEED', 'NON_KILL']

    # Load the CSV file into a DataFrame
    df = pd.read_csv(prediction_csv, header=None, names=['behavior'])


    
    # Define parameters
    # MIN_STALK_DELAY = 32  # Minimum delay for STALK behavior
    # MIN_STALK_TIME = 32   # Minimum time for STALK behavior
    # MIN_KILL_TIME = 32    # Minimum time for KILL behavior
    # MAX_KILL_TIME = 64    # Maximum time for KILL behavior
    # MIN_FEED_DELTA = 32   # Minimum time between KILL and FEED behavior
    # MIN_FEED_DURATION = 32  # Minimum duration of FEED behavior


    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active

    # Set headers for the Excel sheet
    headers = ['A', 'B', 'C', 'D', 'E']
    ws.append(headers)

    # Initialize a 2D array to store pass/fail status of each cell
    pass_fail_status = []

    # Iterate over the behavior data to analyze kill windows
    kill_windows = calculate_duration(df['behavior'], beh_names.index('KILL'))
    for window_duration in kill_windows:
        row_values = []
        row_pass_fail = []  # Store pass/fail status for this row

        # Check for STALK behavior before the start of the kill window
        stalk_before_start = any(df['behavior'].iloc[max(0, i - MIN_STALK_DELAY):i].eq(beh_names.index('STALK')).any() for i in range(window_duration))
        row_values.append(MIN_STALK_TIME if stalk_before_start else '')
        row_pass_fail.append(stalk_before_start)

        # Check the duration of the kill window
        row_values.append(window_duration)
        row_pass_fail.append(MIN_KILL_TIME <= window_duration <= MAX_KILL_TIME)

        # Check for feeding after the kill window
        feed_after_kill = df['behavior'].iloc[window_duration:].eq(beh_names.index('FEED')).any()
        row_values.append(MIN_FEED_DELTA if feed_after_kill else '')
        row_pass_fail.append(feed_after_kill)

        # Check the duration of the feeding window
        feed_duration = calculate_duration(df['behavior'].iloc[window_duration:], beh_names.index('FEED'))
        row_values.append(min(feed_duration) if any(d >= MIN_FEED_DURATION for d in feed_duration) else '')
        row_pass_fail.append(any(d >= MIN_FEED_DURATION for d in feed_duration))

        # Append row values to the Excel sheet
        ws.append(row_values)
        pass_fail_status.append(row_pass_fail)

    # Color cells based on pass/fail status
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=len(kill_windows) + 1, min_col=1, max_col=len(headers))):
        if i < len(pass_fail_status):  # Ensure pass_fail_status has enough rows
            for j, cell in enumerate(row):
                if j < len(pass_fail_status[i]):  # Ensure pass_fail_status has enough columns
                    if not pass_fail_status[i][j]:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        else:
            for cell in row:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")



    # Save the Excel workbook
    fname = os.path.join(output_dir, "kill_window_analysis.xlsx")
    wb.save(fname)
    print(f"Created file {fname}")
# Example usage:


    # df = pd.read_csv(prediction_csv, header=None, names=['behavior'])

    # # Define the behavior labels
    # beh_names = ['unknown', 'STALK', 'KILL', 'KILL_PHASE2', 'FEED', 'NON_KILL']

    # # Map behavior labels to their corresponding names
    # df['behavior_name'] = df['behavior'].map(lambda x: beh_names[x])

    # # Identify sections where the behavior is 'KILL'
    # kill_sections = df[df['behavior_name'] == 'KILL']

    # # Count the number of contiguous sections of 'KILL' behavior
    # total_kill_groupings = 0
    # in_kill_section = False
    # for i in range(len(kill_sections)):
    #     if not in_kill_section:
    #         total_kill_groupings += 1
    #         in_kill_section = True
    #     if i < len(kill_sections) - 1 and kill_sections.index[i] + 1 != kill_sections.index[i+1]:
    #         in_kill_section = False


    # print(f"Total number of kill sections {total_kill_groupings=}")

    # # TODO : return updated csv
    # return prediction_csv

def plot_updated(prediction_dir, config_dir, output_dir):
    config_fp = os.path.join(config_dir, 'dataset_metadata.yaml')
    with open(config_fp) as file:
        metadata = yaml.load(file, Loader=yaml.FullLoader)

    eval_fp = os.path.join(prediction_dir, 'test_eval.yaml')
    with open(eval_fp) as file:
        eval_dict = yaml.load(file, Loader=yaml.FullLoader)


    input_csv = '/home/matthew/AI_Capstone/ai-capstone/BEBE-datasets/F202_15sample_16hz_2hr/FormatData/clip_data/exp999_user2024.csv'
    output_csv = '/home/matthew/AI_Capstone/ai-capstone/BEBE_results_phase2/CRNN/F202_15sample_16hz_2hr_CRNN/F202_15sample_16hz_2hr_CRNN/fold_1/predictions/exp999_user2024.csv'

    input_csv = '/home/matthew/AI_Capstone/ai-capstone/BEBE-datasets/F202_15sample_16hz_2hr/FormatData/clip_data/exp61_user2021.csv'
    output_csv = '/home/matthew/AI_Capstone/ai-capstone/BEBE_results_phase2/harnet/F202_15sample_16hz_2hr_harnet/F202_15sample_16hz_2hr_harnet/fold_1/predictions/exp61_user2021.csv'
    filtered_csv = os.path.join(output_dir, 'filtered.csv')
    filter_model_predictions(output_csv, filtered_csv)
    create_spreadsheet(output_csv, filtered_csv, output_dir)

    num_clusters = len(metadata['label_names']) - 1    
    input_data = pd.read_csv(input_csv, delimiter = ',', header = None).values
    target_fp = os.path.join(output_dir, 'test.png')

    plot_track(data_fp=input_csv,
                predictions_fp=output_csv,
                metadata=metadata,
                num_clusters=num_clusters,
                unsupervised=False,
                eval_dict=eval_dict,
                start_sample=0,
                end_sample=len(input_data) - 1,
                target_fp=target_fp,
                filtered_predictions=filtered_csv)

    print(f"Created plot {target_fp}")
    

def check_dir(input_dir, output_dir, config_dir):
    pass

def print_info():
    """
    Print human-readable criteria used in this run
    :return:
    """
    print(f"Sampling rate: {SAMPLING_RATE} hz")
    print(f"Min. stalk delay: {MIN_STALK_DELAY / SAMPLING_RATE} sec.")
    print(f"Min. stalk duration: {MIN_STALK_TIME / SAMPLING_RATE} sec.")
    print(f"Min. kill time: {MIN_KILL_TIME / SAMPLING_RATE} sec.")
    print(f"Min. feed delta: {MIN_FEED_DELTA / SAMPLING_RATE / 60} min.")
    print(f"Max. feed delta: {MAX_FEED_DELTA / SAMPLING_RATE / 60} min.")
    print(f"Min. feed time: {MIN_FEED_TIME / SAMPLING_RATE / 60} min.")

def main():
    print_info()
    parser = argparse.ArgumentParser(description=summary, formatter_class=argparse.RawTextHelpFormatter)
    # parser.add_argument('--prediction_dir', metavar='DIR', required=True, type=str, help='Directory with outputs from BEBE (should have test_eval.yaml, predictions dir)')
    # parser.add_argument('--output_dir', metavar='DIR', required=True, type=str, default='.', help='Output directory for plots')
    # parser.add_argument('--model_output_dir', metavar='DIR', type=str, default='.', help='Output directory for plots')
    # parser.add_argument('--config_dir', metavar='DIR', required=True, type=str, default='.', help='Directory where clip data/metadata yaml file exists')
    # parser.add_argument('--show-plots', action='store_true', help='Display plots instead of saving to file')
    # parser.add_argument('--quiet', action='store_true', help='Omit summary printouts')

    # args = parser.parse_args()

    # prediction_dir = args.prediction_dir
    # output_dir = args.output_dir
    # config_dir = args.config_dir
    # os.makedirs(output_dir, exist_ok=True)
    # check_dir(directory_path, output_dir, config_dir)

    # plot_updated(prediction_dir, config_dir, output_dir)

    input_csv = '/home/matthew/AI_Capstone/output/BEBE-results/CRNN/both_feeding_8_hz/both_feeding_8_hz_CRNN/fold_1/predictions/exp999_user2024.csv'
    filtered_csv = '/home/matthew/AI_Capstone/output/BEBE-results/CRNN/both_feeding_8_hz/both_feeding_8_hz_CRNN/fold_1/predictions/exp999_user2024_filtered.csv'
    filter_model_predictions(input_csv, filtered_csv)

    print("DONE")


if __name__ == "__main__":
    main()

